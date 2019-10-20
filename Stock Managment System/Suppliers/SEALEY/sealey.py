import openpyxl, os, yaml, csv
import Server.Receive.supplier_ftp as supplier_ftp

#Download file from FTP
def sealey(zero=False):
    if zero:
        print("Zeroing Sealey")
    else:
        print("Starting Sealey")

    supplier_ftp.getFile("sealey")

    with open("config.yml", 'r') as cfg:
        config = yaml.load(cfg, Loader=yaml.FullLoader)

    MAX = config['suppliers']['sealey']['max']
    MIN = config['suppliers']['sealey']['min']

    # Create alter list
    alter = []
    with open('Suppliers/alterlist.csv') as altercsv:
        csvlist = list(csv.reader(altercsv))
        skuind = csvlist[0].index("sealey")
        quaind = skuind + 1
        for row in csvlist[2:]:
            alter.append([row[skuind],row[quaind]])

    print("Loading file...")
    doc = openpyxl.load_workbook('Suppliers/SEALEY/Datacut.xlsx')
    sheet = doc['Export_models_select_files_xls']

    print("Building list")
    vallist = []
    for row in range(2, sheet.max_row + 1):
        sku = str(sheet.cell(row=row, column=1).value+"-SY")
        stock = int(sheet.cell(row=row, column = 5).value)

        # Alter
        if sku in [line[0] for line in alter]:
            stock = alter[[line[0] for line in alter].index(sku)][1]
            
        if stock > MAX:
            stock = MAX
        elif stock < MIN:
            stock = 0

        if zero:
            stock = 0

        vallist.append([sku,stock])

    print("Writing to file")
    with open("Server/Send/StockFiles/sealey.tsv", "w", newline="") as f:
        writer = csv.writer(f, delimiter='\t')
        writer.writerows(vallist)

    print("finished sealey.py")
