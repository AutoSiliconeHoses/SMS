import openpyxl, time, os, zipfile, glob

def decco():
    print("Starting Decco")
    doc = openpyxl.load_workbook('')
    sheet = doc['']

    vallist = []
    for row in range(2, sheet.max_row + 1):
        vallist.append((sheet.cell(row=row, column=1).value+"-SY", sheet.cell(row=row, column = 5).value))

    temptext = ""
    for i in range(0, len(vallist)):
        temptext = temptext + vallist[i][0] + "\t" + str(vallist[i][1]) + "\n"

    with open("Server/Send/StockFiles/decco.tsv", "w") as txtfile:
        txtfile.write(temptext)
        txtfile.close()
    print("Finished Decco")