import openpyxl, os, glob, yaml, sys

def fps_leeds():
    with open("config.yml", 'r') as cfg:
        config = yaml.load(cfg, Loader=yaml.FullLoader)

    MAX = config['suppliers']['fps']['max']

    temptext = ""

    if os.path.isfile('Server/Receive/Emails/FPS_LEEDS.xlsx'):
        for file in glob.glob('Server/Receive/Emails/FPS_LEEDS.xlsx'):
            doc = openpyxl.load_workbook(file)
            sheet = doc.active

            for x in range(1, sheet.max_column+1):
                if sheet.cell(row = 1, column = x).value == "Supplier":
                    supplier = x
                if sheet.cell(row = 1, column = x).value == "Part No":
                    partno = x
                if sheet.cell(row = 1, column = x).value == "Free Stk":
                    freestk = x

            for row in range(2, sheet.max_row + 1):
                suppcode = sheet.cell(row = row, column = partno).value + "-" + sheet.cell(row = row, column = supplier).value[0:3] + "-FPS"
                if sheet.cell(row = row, column = freestk).value == "YES":
                    stk = MAX
                elif sheet.cell(row = row, column = freestk).value == "NO":
                    stk = "0"
                temptext += suppcode + "\t" + str(stk) + "\n"

        with open("Suppliers/FPS/fps_leeds.tsv", "w") as text:
            text.write(temptext)
            text.close()
    else:
        print("Error: FPS_LEEDS file not found")
