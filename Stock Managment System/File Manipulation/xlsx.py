import openpyxl

doc = openpyxl.load_workbook('testxlsx.xlsx')
sheet = doc.get_sheet_by_name('Sheet1')
# val = sheet.cell(row=1, column=1).value
# print(val)

vallist = []
multicell = sheet[sheet.cell(1, 1):sheet.cell(sheet.max_row, sheet.max_column)]
for row in range(1, sheet.max_row + 1):
    vallist.append(sheet.cell(row=row, column=1).value)
    # vallist[vallist.length+1][0]=column=1
    # vallist[vallist.length+1][1]=column=2

    # vallist.append

print(vallist)
